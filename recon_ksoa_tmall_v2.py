"""
KSOA天猫对账 - v2 重构版
核心改变：
1. 匹配链路：doc_no → sys_order → 所有 tmall 订单（原订单号，可能多个用|分隔）
2. 按 doc_no 聚合数据，不再在商品行级别匹配
3. 差异验证：KSOA小计 = 支付宝到账合计 + 聚合结算到账合计
4. 平台费用独立展示，不参与等式验证
5. 交易关闭订单标注但不排除，参与验证
"""
import openpyxl
import xlrd
import os
import json
import re
import datetime
from collections import defaultdict

def json_serializable(obj):
    if isinstance(obj, datetime.datetime):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(obj, datetime.date):
        return obj.strftime('%Y-%m-%d')
    return str(obj)

PATH = r'E:\天猫\对账\2026年2-4月\数据\new'
OUT_DIR = r'D:\Financial check'

# ============================================================
# Step 1: Load all source data
# ============================================================
print('=' * 60)
print('Step 1: Loading source data...')

# 1.1 KSOA
wb_ksoa = xlrd.open_workbook(os.path.join(PATH, 'KSOA零售-天猫.xls'))
ws_ksoa = wb_ksoa.sheet_by_index(0)
ksoa_headers = [ws_ksoa.cell_value(0, c) for c in range(ws_ksoa.ncols)]
ksoa_col = {h: c for c, h in enumerate(ksoa_headers) if h}
print(f'  KSOA: {ws_ksoa.nrows - 1} rows')

# 1.2 万里牛出库退货明细
wb_wln_out = openpyxl.load_workbook(os.path.join(PATH, '万里牛销售出库退货明细.xlsx'), data_only=True)
ws_wln_out = wb_wln_out.active
wln_out_col = {ws_wln_out.cell(row=1, column=c).value: c for c in range(1, ws_wln_out.max_column + 1)
               if ws_wln_out.cell(row=1, column=c).value}
print(f'  万里牛出库: {ws_wln_out.max_row - 1} rows')

# 1.3 万里牛订单明细
wb_wln = openpyxl.load_workbook(os.path.join(PATH, '万里牛订单明细.xlsx'), data_only=True)
ws_wln = wb_wln.active
wln_col = {ws_wln.cell(row=1, column=c).value: c for c in range(1, ws_wln.max_column + 1)
           if ws_wln.cell(row=1, column=c).value}
print(f'  万里牛订单: {ws_wln.max_row - 1} rows')

# 1.4 天猫订单报表
wb_tmall = openpyxl.load_workbook(os.path.join(PATH, '天猫订单明细报表.xlsx'), data_only=True)
ws_tmall = wb_tmall.active
tmall_col = {ws_tmall.cell(row=1, column=c).value: c for c in range(1, ws_tmall.max_column + 1)
             if ws_tmall.cell(row=1, column=c).value}
print(f'  天猫订单: {ws_tmall.max_row - 1} rows')

# 1.5 支付宝交易订单明细
wb_alipay = xlrd.open_workbook(os.path.join(PATH, '支付宝交易订单明细.xls'))
ws_alipay = wb_alipay.sheet_by_index(0)
alipay_headers = [ws_alipay.cell_value(2, c) for c in range(ws_alipay.ncols)]
alipay_col = {h: c for c, h in enumerate(alipay_headers) if h}
print(f'  支付宝交易订单: {ws_alipay.nrows - 3} rows')

# 1.6 支付宝资金流水
wb_fund = openpyxl.load_workbook(os.path.join(PATH, '支付宝资金流水.xlsx'), data_only=True)
ws_fund = wb_fund.active
fund_col = {ws_fund.cell(row=3, column=c).value: c for c in range(1, ws_fund.max_column + 1)
            if ws_fund.cell(row=3, column=c).value}
print(f'  支付宝资金流水: {ws_fund.max_row - 3} rows')

# 1.7 聚合结算
wb_agg = openpyxl.load_workbook(os.path.join(PATH, '聚合结算账户收支明细.xlsx'), data_only=True)
ws_agg = wb_agg.active
agg_col = {ws_agg.cell(row=1, column=c).value: c for c in range(1, ws_agg.max_column + 1)
           if ws_agg.cell(row=1, column=c).value}
print(f'  聚合结算: {ws_agg.max_row - 1} rows')

# ============================================================
# Step 2: Build indexes
# ============================================================
print()
print('=' * 60)
print('Step 2: Building indexes...')

# 2.1 万里牛出库 fallback: doc_no -> sys_order（取第一条）
outbound_fallback = {}
for r in range(2, ws_wln_out.max_row + 1):
    doc = ws_wln_out.cell(row=r, column=wln_out_col.get('出入库单号', 1)).value
    sys_order = ws_wln_out.cell(row=r, column=wln_out_col.get('系统订单号', 24)).value
    if doc and str(doc) not in outbound_fallback:
        outbound_fallback[str(doc)] = str(sys_order) if sys_order else None

# 2.2 万里牛系统单 -> 原订单号（天猫订单号）
# 注意：一个 sys_order 可能有多行，每行对应一个天猫订单（可能重复）
# 聚合为一个去重的订单列表
wln_sys_to_origin = defaultdict(set)
for r in range(2, ws_wln.max_row + 1):
    sys_order = ws_wln.cell(row=r, column=wln_col.get('系统单号', 3)).value
    origin = ws_wln.cell(row=r, column=wln_col.get('原订单号', 27)).value
    if sys_order:
        sys_str = str(sys_order)
        origin_str = str(origin) if origin else ''
        if origin_str:
            if '|' in origin_str:
                for t in origin_str.split('|'):
                    t = t.strip()
                    if t:
                        wln_sys_to_origin[sys_str].add(t)
            else:
                wln_sys_to_origin[sys_str].add(origin_str)
# 转为 list
wln_sys_to_origin = {k: list(v) for k, v in wln_sys_to_origin.items()}

# 2.3 天猫订单索引: order_no -> dict
tmall_index = {}
for r in range(2, ws_tmall.max_row + 1):
    order = ws_tmall.cell(row=r, column=tmall_col.get('订单编号', 1)).value
    if order:
        tmall_index[str(order)] = {
            'status': ws_tmall.cell(row=r, column=tmall_col.get('订单状态', 9)).value,
            'buyer_paid': ws_tmall.cell(row=r, column=tmall_col.get('买家实际支付金额', 7)).value,
            'create_time': ws_tmall.cell(row=r, column=tmall_col.get('订单创建时间', 12)).value,
            'pay_time': ws_tmall.cell(row=r, column=tmall_col.get('订单付款时间', 13)).value,
            'confirm_time': ws_tmall.cell(row=r, column=tmall_col.get('确认收货时间', 23)).value,
            'merchant_received': ws_tmall.cell(row=r, column=tmall_col.get('打款商家金额', 24)).value,
        }

# 2.4 支付宝交易订单索引: tmall_order -> 商家实收
alipay_order_index = {}
for r in range(3, ws_alipay.nrows):
    merchant_order = ws_alipay.cell_value(r, alipay_col.get('商户订单号', 4))
    if merchant_order and str(merchant_order).startswith('T200P'):
        tmall = str(merchant_order)[5:]
        alipay_order_index[tmall] = ws_alipay.cell_value(r, alipay_col.get('商家实收(元)', 13))

# 2.5 支付宝资金流水索引: tmall_order -> list of {type, income, expense, remark, biz_desc}
fund_flow_index = defaultdict(list)
for r in range(4, ws_fund.max_row + 1):
    base_order = ws_fund.cell(row=r, column=fund_col.get('业务基础订单号', 15)).value
    biz_type = ws_fund.cell(row=r, column=fund_col.get('账务类型', 6)).value
    income_val = ws_fund.cell(row=r, column=fund_col.get('收入（+元）', 7)).value
    expense_val = ws_fund.cell(row=r, column=fund_col.get('支出（-元）', 8)).value
    remark = ws_fund.cell(row=r, column=fund_col.get('备注', 14)).value
    biz_desc = ws_fund.cell(row=r, column=fund_col.get('业务描述', 18)).value
    try:
        income = float(income_val) if income_val and str(income_val).strip() else 0
    except:
        income = 0
    try:
        expense = float(expense_val) if expense_val and str(expense_val).strip() else 0
    except:
        expense = 0
    if base_order:
        fund_flow_index[str(base_order)].append({
            'biz_type': biz_type,
            'income': income,
            'expense': abs(expense),
            'remark': remark or '',
            'biz_desc': biz_desc or '',
        })

# 2.6 聚合结算索引: tmall_order -> {income_total, expense_total, detail}
agg_index = defaultdict(lambda: {'income': 0, 'expense': 0, 'detail': []})
for r in range(2, ws_agg.max_row + 1):
    order = ws_agg.cell(row=r, column=agg_col.get('淘宝订单编号', 3)).value
    entry_type = ws_agg.cell(row=r, column=agg_col.get('入账类型', 4)).value
    income_val = ws_agg.cell(row=r, column=agg_col.get('收入金额（元）', 5)).value
    expense_val = ws_agg.cell(row=r, column=agg_col.get('支出金额', 6)).value
    biz_desc = ws_agg.cell(row=r, column=agg_col.get('业务描述', 7)).value
    remark = ws_agg.cell(row=r, column=agg_col.get('备注', 8)).value
    try:
        income = float(str(income_val).replace(',', '')) if income_val and str(income_val).strip() else 0
    except:
        income = 0
    try:
        expense = float(str(expense_val).replace(',', '')) if expense_val and str(expense_val).strip() else 0
    except:
        expense = 0
    if order:
        o = str(order)
        agg_index[o]['detail'].append({
            'type': entry_type,
            'income': income,
            'expense': expense,
            'biz_desc': biz_desc or '',
            'remark': remark or '',
        })
        if entry_type in ['交易收款', '扣款退回']:
            agg_index[o]['income'] += income
        elif entry_type in ['扣款', '交易退款(售后)']:
            agg_index[o]['expense'] += expense

print(f'  Outbound fallback: {len(outbound_fallback)}')
print(f'  Wanliuniu sys->origin: {len(wln_sys_to_origin)}')
print(f'  Tmall orders: {len(tmall_index)}')
print(f'  Alipay order index: {len(alipay_order_index)}')
print(f'  Fund flow: {len(fund_flow_index)} orders')
print(f'  Aggregation: {len(agg_index)} orders')

# ============================================================
# Step 3: Group KSOA rows by doc_no
# ============================================================
print()
print('=' * 60)
print('Step 3: Grouping KSOA rows by doc_no...')

# 按 doc_no 分组 KSOA 数据
doc_groups = defaultdict(list)  # doc_no -> list of row data
sales_records = []
return_records = []

for r in range(1, ws_ksoa.nrows):
    doc_no = ws_ksoa.cell_value(r, ksoa_col.get('单据编号', 0))
    qty = ws_ksoa.cell_value(r, ksoa_col.get('数量', 12))
    real_amount = ws_ksoa.cell_value(r, ksoa_col.get('实收金额', 20))
    product = ws_ksoa.cell_value(r, ksoa_col.get('商品名称', 6))
    date = ws_ksoa.cell_value(r, ksoa_col.get('日期', 3))
    if not doc_no:
        continue

    record = {
        'ksoa_doc': str(doc_no),
        'date': str(date) if date else '',
        'product': product,
        'qty': qty,
        'ksoa_amount': real_amount,
        'wln_sys_order': None,
    }

    is_return = (qty < 0) if qty else False

    # 通过 outbound_fallback + wln_sys_to_origin 查找 tmall 订单
    sys_order = outbound_fallback.get(str(doc_no))
    if sys_order:
        tmall_list = wln_sys_to_origin.get(sys_order, [])
        record['wln_sys_order'] = sys_order
        record['tmall_orders'] = tmall_list  # 可能多个，用 | 分隔
    else:
        record['tmall_orders'] = []

    if is_return:
        return_records.append(record)
    else:
        sales_records.append(record)
        doc_groups[str(doc_no)].append(record)

print(f'  Sales docs: {len(doc_groups)}')
print(f'  Return records: {len(return_records)}')

# 统计匹配情况
docs_with_tmall = sum(1 for doc, rows in doc_groups.items()
                      if rows[0].get('tmall_orders'))
print(f'  Docs matched to tmall: {docs_with_tmall}')

# ============================================================
# Step 4: Build output - per doc aggregation
# ============================================================
print()
print('=' * 60)
print('Step 4: Building output (doc-level aggregation)...')

def build_expense_str(detail):
    """Build fund expense detail string."""
    parts = []
    for f in detail:
        desc = f.get('biz_desc', '') or f.get('remark', '') or ''
        parts.append(f"{f['biz_type']}: {desc} (¥{f['expense']:.2f})")
    return '; '.join(parts) if parts else ''

def build_agg_str(detail):
    """Build aggregation settlement detail string."""
    parts = []
    for f in detail:
        desc = f.get('biz_desc', '') or f.get('remark', '') or ''
        amt = f['income'] if f['income'] > 0 else -f['expense']
        parts.append(f"{f['type']}: {desc} (¥{abs(amt):.2f})")
    return '; '.join(parts) if parts else ''

sales_detail = []

for doc, rows in doc_groups.items():
    # KSOA 小计（doc 下所有行金额之和）
    ksoa_sum = sum(float(r['ksoa_amount']) for r in rows if r['ksoa_amount'])

    # 获取关联的 tmall 订单列表
    first_row = rows[0]
    tmall_orders = first_row.get('tmall_orders', [])

    # 如果没有匹配到 tmall，标记为未匹配
    if not tmall_orders:
        row = {
            'KSOA单据编号': doc,
            '日期': first_row['date'],
            '商品名称': first_row['product'],
            '数量': first_row['qty'],
            'KSOA实收金额': first_row['ksoa_amount'],
            '万里牛系统订单号': first_row.get('wln_sys_order'),
            '天猫订单号': None,
            '天猫订单状态': None,
            '买家实付金额': None,
            '订单创建时间': None,
            '付款时间': None,
            '确认收货时间': None,
            '支付宝商家实收': None,
            '支付宝到账': None,
            '聚合结算到账': None,
            '聚合结算支出金额': None,
            '平台费用支出': None,
            '平台费用支出明细': '',
            '聚合结算明细': '',
            'KSOA小计': round(ksoa_sum, 2),
            '等式验证': None,
            '等式差异': None,
            '差异原因细化': '未匹配到万里牛出库记录',
        }
        sales_detail.append(row)
        continue

    # 合并所有 tmall 订单的数据
    all_tmall_str = '; '.join(sorted(tmall_orders))

    total_alipay_inc = 0.0   # 支付宝到账合计
    total_agg_inc = 0.0      # 聚合结算到账合计
    total_fund_exp = 0.0     # 支付宝平台费用支出合计
    total_agg_exp = 0.0      # 聚合结算支出合计
    total_buyer_paid = 0.0   # 买家实付合计
    total_merchant_received = 0.0  # 支付宝商家实收合计
    fund_exp_detail_list = []
    agg_detail_list = []
    tmall_statuses = set()
    order_create_time = None
    order_pay_time = None
    order_confirm_time = None
    has_closed = False  # 是否有交易关闭订单

    for tm in tmall_orders:
        # 天猫订单信息
        tm_info = tmall_index.get(tm, {})
        if tm_info.get('buyer_paid'):
            total_buyer_paid += float(tm_info['buyer_paid'])
        if tm_info.get('status'):
            tmall_statuses.add(tm_info['status'])
            if tm_info['status'] == '交易关闭':
                has_closed = True
        if tm_info.get('create_time') and not order_create_time:
            order_create_time = tm_info['create_time']
        if tm_info.get('pay_time') and not order_pay_time:
            order_pay_time = tm_info['pay_time']
        if tm_info.get('confirm_time') and not order_confirm_time:
            order_confirm_time = tm_info['confirm_time']
        if tm_info.get('merchant_received'):
            try:
                val = str(tm_info['merchant_received']).replace('¥', '').replace('元', '').replace(',', '').strip()
                total_merchant_received += float(val)
            except:
                pass

        # 支付宝到账（在线支付）
        for f in fund_flow_index.get(tm, []):
            if f['biz_type'] == '在线支付' and f['income'] > 0:
                total_alipay_inc += f['income']
            # 所有非在线支付的支出都计入平台费用
            if f['expense'] > 0 and f['biz_type'] != '在线支付':
                total_fund_exp += f['expense']
                fund_exp_detail_list.append(f)

        # 聚合结算
        agg = agg_index.get(tm, {})
        total_agg_inc += agg.get('income', 0)
        total_agg_exp += agg.get('expense', 0)
        agg_detail_list.extend(agg.get('detail', []))

    total_fund_exp = round(total_fund_exp, 2)
    total_agg_exp = round(total_agg_exp, 2)
    total_platform_fee = round(total_fund_exp + total_agg_exp, 2)
    total_income = round(total_alipay_inc + total_agg_inc, 2)
    net_received = round(ksoa_sum - total_platform_fee, 2)
    diff = round(ksoa_sum - total_income, 2)

    # 等式验证
    if abs(diff) <= 0.01:
        formula_check = f'{ksoa_sum:.2f} = {total_alipay_inc:.2f} + {total_agg_inc:.2f}'
        diff_reason = '无差异'
    else:
        formula_check = f'{ksoa_sum:.2f} ≠ {total_alipay_inc:.2f} + {total_agg_inc:.2f}'
        reasons = []
        if abs(diff) > 0.01:
            reasons.append(f'等式差异¥{diff:.2f}')
        diff_reason = '; '.join(reasons) if reasons else ''

    # 订单状态标注
    if has_closed:
        status_note = '含关闭订单'
    else:
        status_note = '正常'

    row = {
        'KSOA单据编号': doc,
        '日期': first_row['date'],
        '商品名称': first_row['product'],
        '数量': first_row['qty'],
        'KSOA实收金额': first_row['ksoa_amount'],
        '万里牛系统订单号': first_row.get('wln_sys_order'),
        '天猫订单号': all_tmall_str,
        '天猫订单状态': '; '.join(sorted(tmall_statuses)),
        '订单状态标注': status_note,
        '买家实付金额': round(total_buyer_paid, 2) if total_buyer_paid else None,
        '订单创建时间': order_create_time,
        '付款时间': order_pay_time,
        '确认收货时间': order_confirm_time,
        '支付宝商家实收': round(total_merchant_received, 2) if total_merchant_received else None,
        '支付宝到账': round(total_alipay_inc, 2) if total_alipay_inc else None,
        '聚合结算到账': round(total_agg_inc, 2) if total_agg_inc else None,
        '聚合结算支出金额': round(total_agg_exp, 2) if total_agg_exp else None,
        '平台费用支出': -total_platform_fee if total_platform_fee else None,
        '平台费用支出明细': build_expense_str(fund_exp_detail_list),
        '聚合结算明细': build_agg_str(agg_detail_list),
        'KSOA小计': round(ksoa_sum, 2),
        '等式验证': formula_check,
        '等式差异': round(diff, 2) if abs(diff) > 0.01 else None,
        '净收': net_received if total_platform_fee else None,
        '差异原因细化': diff_reason,
    }
    sales_detail.append(row)

# 退货记录
return_detail = []
for rec in return_records:
    doc = rec['ksoa_doc']
    sys_order = outbound_fallback.get(doc)
    tmall_orders = wln_sys_to_origin.get(sys_order, []) if sys_order else []
    all_tmall_str = '; '.join(sorted(tmall_orders)) if tmall_orders else None

    # 获取第一个 tmall 的状态
    tmall_status = None
    agg_refund = 0.0
    if tmall_orders:
        first_tm = tmall_orders[0]
        tm_info = tmall_index.get(first_tm, {})
        tmall_status = tm_info.get('status')
        agg = agg_index.get(first_tm, {})
        agg_refund = agg.get('expense', 0)

    row = {
        'KSOA单据编号': doc,
        '日期': rec['date'],
        '商品名称': rec['product'],
        '退货数量': abs(rec['qty']),
        'KSOA实收金额': rec['ksoa_amount'],
        '万里牛系统订单号': sys_order,
        '天猫订单号': all_tmall_str,
        '天猫订单状态': tmall_status,
        '聚合结算退款': round(agg_refund, 2) if agg_refund else None,
        '对账差异原因': '' if tmall_orders else '未匹配到万里牛出库记录',
    }
    return_detail.append(row)

# Save JSON
output = {
    'sales_detail': sales_detail,
    'return_detail': return_detail,
}
json_path = os.path.join(OUT_DIR, 'KSOA天猫对账_v2.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2, default=json_serializable)
print(f'JSON saved: {json_path}')

# ============================================================
# Summary stats
# ============================================================
def sf(v):
    try: return float(v) if v else 0
    except: return 0

ksoa_total = sum(sf(r['KSOA小计']) for r in sales_detail)
alipay_total = sum(sf(r.get('支付宝到账')) for r in sales_detail if r.get('支付宝到账'))
agg_total = sum(sf(r.get('聚合结算到账')) for r in sales_detail if r.get('聚合结算到账'))
fee_total = sum(abs(sf(r.get('平台费用支出'))) for r in sales_detail if r.get('平台费用支出'))
diff_count = sum(1 for r in sales_detail if r.get('等式差异') and r.get('等式差异') != 0)
unmatched = sum(1 for r in sales_detail if r.get('差异原因细化') == '未匹配到万里牛出库记录')
has_closed = sum(1 for r in sales_detail if r.get('订单状态标注') == '含关闭订单')

print()
print('=== Summary ===')
print(f'Sales docs: {len(sales_detail)}')
print(f'Return records: {len(return_detail)}')
print(f'KSOA total: {ksoa_total:,.2f}')
print(f'Alipay income: {alipay_total:,.2f}')
print(f'Agg income: {agg_total:,.2f}')
print(f'Total income: {alipay_total + agg_total:,.2f}')
print(f'Platform fee total: {fee_total:,.2f}')
print(f'Diff docs (等式差异≠0): {diff_count}')
print(f'Unmatched: {unmatched}')
print(f'含关闭订单: {has_closed}')