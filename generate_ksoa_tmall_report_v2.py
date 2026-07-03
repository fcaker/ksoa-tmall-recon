"""
KSOA天猫对账 - Excel报告生成器 v2
基于 doc 维度的等式验证结果
"""
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8')

print('Generating KSOA天猫对账 Excel report v2...')

with open(r'D:\Financial check\KSOA天猫对账_v2.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = openpyxl.Workbook()
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
title_font = Font(bold=True, size=14)
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
orange_fill = PatternFill(start_color='FFE4C4', end_color='FFE4C4', fill_type='solid')

def sf(v):
    try: return float(v) if v else 0
    except: return 0

sales = data['sales_detail']
returns = data['return_detail']

# ============================================================
# Sheet 1: 核对汇总
# ============================================================
ws = wb.active
ws.title = '核对汇总'
ws['A1'] = 'KSOA天猫对账核对结果汇总'
ws['A1'].font = title_font
ws.merge_cells('A1:D1')

ksoa_total = sum(sf(r.get('KSOA小计')) for r in sales)
alipay_inc = sum(sf(r.get('支付宝到账')) for r in sales if r.get('支付宝到账'))
agg_inc = sum(sf(r.get('聚合结算到账')) for r in sales if r.get('聚合结算到账'))
fee_exp = sum(abs(sf(r.get('平台费用支出'))) for r in sales if r.get('平台费用支出'))
diff_count = sum(1 for r in sales if r.get('等式差异') and r.get('等式差异') != 0)
unmatched = sum(1 for r in sales if r.get('差异原因细化') == '未匹配到万里牛出库记录')
has_closed = sum(1 for r in sales if r.get('订单状态标注') == '含关闭订单')
no_diff = len(sales) - diff_count - unmatched

summary = [
    ['项目', '数值', '', ''],
    ['KSOA单据数', len(sales), '', ''],
    ['有等式差异', diff_count, '', ''],
    ['无差异（等式成立）', no_diff, '', ''],
    ['含关闭订单', has_closed, '', ''],
    ['未匹配', unmatched, '', ''],
    ['', '', '', ''],
    ['KSOA实收金额合计', f"¥{ksoa_total:,.2f}", '', ''],
    ['支付宝到账合计', f"¥{alipay_inc:,.2f}", '', ''],
    ['聚合结算到账合计', f"¥{agg_inc:,.2f}", '', ''],
    ['合计到账', f"¥{alipay_inc + agg_inc:,.2f}", '', ''],
    ['平台费用支出合计', f"¥{fee_exp:,.2f}", '', ''],
    ['', '', '', ''],
    ['等式说明', 'KSOA小计 = 支付宝到账合计 + 聚合结算到账合计', '', ''],
    ['净收计算', 'KSOA小计 - 平台费用支出 = 净收', '', ''],
]

for row_idx, row_data in enumerate(summary, start=3):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        if row_idx == 3:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

for col in ['A', 'B', 'C', 'D']:
    ws.column_dimensions[col].width = 30

# ============================================================
# Sheet 2: 销售订单核对明细
# ============================================================
ws_sales = wb.create_sheet('销售订单核对明细')

headers = [
    'KSOA单据编号', '日期', '商品名称', '数量', 'KSOA实收金额',
    '万里牛系统订单号', '天猫订单号', '天猫订单状态', '订单状态标注',
    '买家实付金额', '订单创建时间', '付款时间', '确认收货时间',
    '支付宝商家实收', '支付宝到账', '聚合结算到账', '聚合结算支出金额',
    '平台费用支出', '平台费用支出明细', '聚合结算明细',
    'KSOA小计', '等式验证', '等式差异', '净收', '差异原因细化'
]

for col_idx, h in enumerate(headers, 1):
    cell = ws_sales.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for row_idx, rec in enumerate(sales, 2):
    diff_reason = rec.get('差异原因细化', '')
    is_closed = rec.get('订单状态标注') == '含关闭订单'
    has_diff = rec.get('等式差异') and rec.get('等式差异') != 0
    is_unmatched = diff_reason == '未匹配到万里牛出库记录'

    values = [
        rec.get('KSOA单据编号'),
        rec.get('日期'),
        rec.get('商品名称'),
        rec.get('数量'),
        rec.get('KSOA实收金额'),
        rec.get('万里牛系统订单号'),
        rec.get('天猫订单号'),
        rec.get('天猫订单状态'),
        rec.get('订单状态标注'),
        rec.get('买家实付金额'),
        rec.get('订单创建时间'),
        rec.get('付款时间'),
        rec.get('确认收货时间'),
        rec.get('支付宝商家实收'),
        rec.get('支付宝到账'),
        rec.get('聚合结算到账'),
        rec.get('聚合结算支出金额'),
        rec.get('平台费用支出'),
        rec.get('平台费用支出明细'),
        rec.get('聚合结算明细'),
        rec.get('KSOA小计'),
        rec.get('等式验证'),
        rec.get('等式差异'),
        rec.get('净收'),
        rec.get('差异原因细化'),
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws_sales.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        # 颜色优先级：未匹配 > 有差异 > 含关闭 > 正常
        if is_unmatched:
            cell.fill = yellow_fill
        elif has_diff:
            cell.fill = red_fill
        elif is_closed:
            cell.fill = orange_fill

col_widths = [20, 12, 30, 8, 12, 20, 35, 12, 12, 12, 18, 18, 18, 15, 12, 12, 12, 12, 50, 50, 12, 30, 12, 12, 50]
for col_idx, w in enumerate(col_widths, 1):
    ws_sales.column_dimensions[get_column_letter(col_idx)].width = w
ws_sales.row_dimensions[1].height = 30

# ============================================================
# Sheet 3: 退货订单核对明细
# ============================================================
ws_ret = wb.create_sheet('退货订单核对明细')

ret_headers = ['KSOA单据编号', '日期', '商品名称', '退货数量', 'KSOA实收金额',
               '万里牛系统订单号', '天猫订单号', '天猫订单状态', '聚合结算退款', '对账差异原因']

for col_idx, h in enumerate(ret_headers, 1):
    cell = ws_ret.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = border

for row_idx, rec in enumerate(returns, 2):
    is_unmatched = rec.get('对账差异原因') == '未匹配到万里牛出库记录'
    values = [
        rec.get('KSOA单据编号'),
        rec.get('日期'),
        rec.get('商品名称'),
        rec.get('退货数量'),
        rec.get('KSOA实收金额'),
        rec.get('万里牛系统订单号'),
        rec.get('天猫订单号'),
        rec.get('天猫订单状态'),
        rec.get('聚合结算退款'),
        rec.get('对账差异原因'),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws_ret.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        if is_unmatched:
            cell.fill = yellow_fill

ret_widths = [20, 12, 30, 10, 12, 20, 35, 12, 15, 30]
for col_idx, w in enumerate(ret_widths, 1):
    ws_ret.column_dimensions[get_column_letter(col_idx)].width = w

# ============================================================
# Sheet 4: 差异明细（有等式差异的订单）
# ============================================================
ws_diff = wb.create_sheet('差异明细')

diff_headers = ['KSOA单据编号', '日期', '天猫订单号', '天猫订单状态', '订单状态标注',
                '买家实付金额', '支付宝到账', '聚合结算到账', 'KSOA小计',
                '等式验证', '等式差异', '平台费用支出', '净收', '差异原因细化']
for col_idx, h in enumerate(diff_headers, 1):
    cell = ws_diff.cell(row=1, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

row_idx = 2
for rec in sales:
    if rec.get('等式差异') and rec.get('等式差异') != 0:
        values = [
            rec.get('KSOA单据编号'),
            rec.get('日期'),
            rec.get('天猫订单号'),
            rec.get('天猫订单状态'),
            rec.get('订单状态标注'),
            rec.get('买家实付金额'),
            rec.get('支付宝到账'),
            rec.get('聚合结算到账'),
            rec.get('KSOA小计'),
            rec.get('等式验证'),
            rec.get('等式差异'),
            rec.get('平台费用支出'),
            rec.get('净收'),
            rec.get('差异原因细化'),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_diff.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if rec.get('订单状态标注') == '含关闭订单':
                cell.fill = orange_fill
            else:
                cell.fill = red_fill
        row_idx += 1

diff_widths = [20, 12, 35, 12, 12, 12, 12, 12, 12, 30, 12, 12, 12, 50]
for col_idx, w in enumerate(diff_widths, 1):
    ws_diff.column_dimensions[get_column_letter(col_idx)].width = w

# ============================================================
# Sheet 5: 交易关闭订单
# ============================================================
ws_closed = wb.create_sheet('交易关闭订单')
closed = [r for r in sales if r.get('订单状态标注') == '含关闭订单']

ws_closed['A1'] = f'交易关闭订单明细（{len(closed)}个KSOA单据）'
ws_closed['A1'].font = title_font
ws_closed.merge_cells('A1:F1')

closed_headers = ['KSOA单据编号', '日期', '天猫订单号', '天猫订单状态', 'KSOA小计', '支付宝到账', '聚合结算到账', '等式差异', '平台费用支出', '差异原因']
for col_idx, h in enumerate(closed_headers, 1):
    cell = ws_closed.cell(row=3, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center')

for row_idx, rec in enumerate(closed, 4):
    values = [
        rec.get('KSOA单据编号'),
        rec.get('日期'),
        rec.get('天猫订单号'),
        rec.get('天猫订单状态'),
        rec.get('KSOA小计'),
        rec.get('支付宝到账'),
        rec.get('聚合结算到账'),
        rec.get('等式差异'),
        rec.get('平台费用支出'),
        rec.get('差异原因细化'),
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws_closed.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        cell.fill = orange_fill

closed_widths = [20, 12, 35, 12, 12, 12, 12, 12, 12, 50]
for col_idx, w in enumerate(closed_widths, 1):
    ws_closed.column_dimensions[get_column_letter(col_idx)].width = w

# ============================================================
# Sheet 6: 未匹配订单
# ============================================================
ws_unmatch = wb.create_sheet('未匹配订单')
unmatch_sales = [r for r in sales if r.get('差异原因细化') == '未匹配到万里牛出库记录']
unmatch_returns = [r for r in returns if r.get('对账差异原因') == '未匹配到万里牛出库记录']

ws_unmatch['A1'] = f'未匹配订单（销售:{len(unmatch_sales)}笔，退货:{len(unmatch_returns)}笔）'
ws_unmatch['A1'].font = title_font
ws_unmatch.merge_cells('A1:F1')

unmatch_headers = ['KSOA单据编号', '日期', '商品名称', '数量', 'KSOA实收金额', '类型']
for col_idx, h in enumerate(unmatch_headers, 1):
    cell = ws_unmatch.cell(row=3, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border

row_idx = 4
for rec in unmatch_sales:
    for col_idx, val in enumerate([rec.get('KSOA单据编号'), rec.get('日期'),
                                   rec.get('商品名称'), rec.get('数量'),
                                   rec.get('KSOA实收金额'), '销售'], 1):
        ws_unmatch.cell(row=row_idx, column=col_idx, value=val).border = border
    row_idx += 1
for rec in unmatch_returns:
    for col_idx, val in enumerate([rec.get('KSOA单据编号'), rec.get('日期'),
                                   rec.get('商品名称'), rec.get('退货数量'),
                                   rec.get('KSOA实收金额'), '退货'], 1):
        ws_unmatch.cell(row=row_idx, column=col_idx, value=val).border = border
    row_idx += 1

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws_unmatch.column_dimensions[col].width = 20

# ============================================================
# Sheet 7: 字段说明
# ============================================================
ws_field = wb.create_sheet('字段说明')

ws_field['A1'] = '字段说明 - 数据来源与计算公式'
ws_field['A1'].font = title_font
ws_field.merge_cells('A1:E1')

field_headers = ['字段名', '源文件', 'Sheet/行', '列', '说明']
for col_idx, h in enumerate(field_headers, 1):
    cell = ws_field.cell(row=3, column=col_idx, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

field_data = [
    ['KSOA单据编号', 'KSOA零售-天猫.xls', 'Row0', 'col0 (A)', 'KSOA销售单据编号，主键'],
    ['日期', 'KSOA零售-天猫.xls', 'Row0', 'col3 (D)', '业务日期'],
    ['商品名称', 'KSOA零售-天猫.xls', 'Row0', 'col6 (G)', '商品名称（展示第一条）'],
    ['数量', 'KSOA零售-天猫.xls', 'Row0', 'col12 (M)', '销售数量（展示第一条）'],
    ['KSOA实收金额', 'KSOA零售-天猫.xls', 'Row0', 'col20 (U)', '该行实收金额（展示第一条）'],
    ['万里牛系统订单号', '万里牛销售出库退货明细.xlsx', 'Row1', 'col24 (X)', '通过 doc_no → outbound_fallback → wln_sys_to_origin 获取'],
    ['天猫订单号', '万里牛订单明细.xlsx', 'Row1', 'col27 (原订单号)', '一个KSOA单据可能对应多个天猫订单，用分号分隔'],
    ['天猫订单状态', '天猫订单明细报表.xlsx', 'Row1', 'col9 (J)', '按天猫订单号匹配，多订单时用分号拼接'],
    ['买家实付金额', '天猫订单明细报表.xlsx', 'Row1', 'col7 (H)', '所有关联天猫订单的买家实付金额合计'],
    ['支付宝到账', '支付宝资金流水.xlsx', 'Row3', 'col7 (G)', '账务类型="在线支付"的收入合计'],
    ['聚合结算到账', '聚合结算账户收支明细.xlsx', 'Row1', 'col5 (E)', '入账类型="交易收款"+"扣款退回"的收入合计'],
    ['聚合结算支出金额', '聚合结算账户收支明细.xlsx', 'Row1', 'col6 (F)', '入账类型="交易退款(售后)"+"扣款"的支出合计'],
    ['平台费用支出', '支付宝资金流水.xlsx', 'Row3', 'col8 (H)', '账务类型非"在线支付"的支出合计（负数）'],
    ['KSOA小计', '汇总计算', '-', '-', '同一KSOA单据编号下所有行实收金额之和'],
    ['等式验证', '汇总计算', '-', '-', 'KSOA小计 = 支付宝到账 + 聚合结算到账'],
    ['等式差异', '汇总计算', '-', '-', 'KSOA小计 - (支付宝到账 + 聚合结算到账)，绝对值>0.01时列出'],
    ['净收', '汇总计算', '-', '-', 'KSOA小计 + 平台费用支出（平台费用为负数）'],
    ['差异原因细化', '汇总计算', '-', '-', '等式差异金额说明，或"无差异"'],
    ['订单状态标注', '汇总计算', '-', '-', '"含关闭订单"：任一关联天猫订单状态为交易关闭；否则"正常"'],
]

for row_idx, row_data in enumerate(field_data, 4):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_field.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws_field.column_dimensions['A'].width = 22
ws_field.column_dimensions['B'].width = 30
ws_field.column_dimensions['C'].width = 12
ws_field.column_dimensions['D'].width = 18
ws_field.column_dimensions['E'].width = 45

# ============================================================
# Save
# ============================================================
outpath = r'D:\Financial check\KSOA天猫对账_v2.xlsx'
wb.save(outpath)
print(f'Report saved: {outpath}')

print()
print('=== KSOA天猫对账 v2 Summary ===')
print(f'Sales docs: {len(sales)}')
print(f'Return records: {len(returns)}')
print(f'KSOA total: ¥{ksoa_total:,.2f}')
print(f'Alipay income: ¥{alipay_inc:,.2f}')
print(f'Agg income: ¥{agg_inc:,.2f}')
print(f'Total income: ¥{alipay_inc + agg_inc:,.2f}')
print(f'Platform fee: ¥{fee_exp:,.2f}')
print(f'Diff docs: {diff_count}')
print(f'含关闭订单: {has_closed}')
print(f'Unmatched: {unmatched}')
print(f'No diff (等式成立): {no_diff}')